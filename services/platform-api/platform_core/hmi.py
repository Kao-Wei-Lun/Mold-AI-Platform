import hashlib
import io
import json
import uuid
from dataclasses import dataclass

from django.conf import settings
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageDraw, ImageOps, UnidentifiedImageError

from .models import (
    Artifact,
    ArtifactVersion,
    AuditEvent,
    HMICorrectionDecision,
    HMIExport,
    HMIExtractedField,
    HMIExtraction,
    HMIProfileVersion,
)

PROFILE_KEY = "demo-generic-injection"
PROFILE_VERSION = "1.0"
EXTRACTOR_VERSION = "seven-segment-profile@1.0.0"
TEMPLATE_VERSION = "reviewed-parameters@1.0.0"
IMAGE_SIZE = (800, 500)
REGION_SIZE = (280, 50)
CONFIDENCE_THRESHOLD = 0.90
MAX_IMAGE_PIXELS = 20_000_000


class HMIValidationError(ValueError):
    def __init__(self, code: str, user_message: str) -> None:
        super().__init__(user_message)
        self.code = code
        self.user_message = user_message


@dataclass(frozen=True)
class FieldSpec:
    code: str
    label: str
    unit: str
    minimum: float
    maximum: float
    region: tuple[float, float, float, float]
    demo_value: float


FIELD_SPECS = (
    FieldSpec(
        "injection_pressure_mpa",
        "Injection Pressure",
        "MPa",
        20,
        300,
        (0.55, 0.16, 0.35, 0.10),
        120,
    ),
    FieldSpec(
        "injection_speed_mm_s",
        "Injection Speed",
        "mm/s",
        1,
        500,
        (0.55, 0.34, 0.35, 0.10),
        45,
    ),
    FieldSpec(
        "holding_pressure_mpa",
        "Holding Pressure",
        "MPa",
        5,
        250,
        (0.55, 0.52, 0.35, 0.10),
        55,
    ),
    FieldSpec(
        "cooling_time_s",
        "Cooling Time",
        "s",
        0,
        300,
        (0.55, 0.70, 0.35, 0.10),
        18,
    ),
)


def _profile_specs_payload() -> list[dict[str, object]]:
    return [
        {
            "code": spec.code,
            "label": spec.label,
            "unit": spec.unit,
            "minimum": spec.minimum,
            "maximum": spec.maximum,
            "region": list(spec.region),
        }
        for spec in FIELD_SPECS
    ]


def get_published_hmi_profile() -> HMIProfileVersion:
    specs = _profile_specs_payload()
    checksum = _sha256(json.dumps(specs, sort_keys=True).encode())
    profile = HMIProfileVersion.objects.filter(
        profile_key=PROFILE_KEY, status=HMIProfileVersion.Status.PUBLISHED
    ).first()
    if profile is None and not HMIProfileVersion.objects.filter(profile_key=PROFILE_KEY).exists():
        profile = HMIProfileVersion.objects.create(
            profile_key=PROFILE_KEY,
            version=PROFILE_VERSION,
            status=HMIProfileVersion.Status.PUBLISHED,
            field_specs=specs,
            profile_checksum=checksum,
            change_summary="Synthetic fixed-layout seven-segment Demo profile",
            created_by="system:demo",
            published_by="system:demo",
            published_at=timezone.now(),
        )
    if profile is None:
        raise HMIValidationError("HMI_PROFILE_NOT_PUBLISHED", "No published HMI profile exists.")
    if profile.profile_checksum != checksum:
        raise HMIValidationError(
            "HMI_PROFILE_CHECKSUM_MISMATCH", "The published HMI profile checksum differs."
        )
    return profile


SEGMENTS = {
    "0": "abcdef",
    "1": "bc",
    "2": "abdeg",
    "3": "abcdg",
    "4": "bcfg",
    "5": "acdfg",
    "6": "acdefg",
    "7": "abc",
    "8": "abcdefg",
    "9": "abcdfg",
}
DIGIT_ORIGINS = (15, 65, 115, 185)
SEGMENT_POINTS = {
    "a": (15, 3),
    "b": (28, 12),
    "c": (28, 32),
    "d": (15, 41),
    "e": (2, 32),
    "f": (2, 12),
    "g": (15, 22),
}
SEGMENT_RECTS = {
    "a": (5, 1, 25, 5),
    "b": (25, 4, 29, 20),
    "c": (25, 24, 29, 40),
    "d": (5, 39, 25, 43),
    "e": (1, 24, 5, 40),
    "f": (1, 4, 5, 20),
    "g": (5, 20, 25, 24),
}


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _draw_number(
    draw: ImageDraw.ImageDraw, region_box: tuple[int, int, int, int], value: float, *, dim: bool
) -> None:
    left, top, _, _ = region_box
    text = f"{int(value):03d}0"
    color = (115, 150, 118) if dim else (72, 244, 126)
    inactive = (8, 18, 14)
    for index, digit in enumerate(text):
        origin_x = left + DIGIT_ORIGINS[index]
        for segment, rect in SEGMENT_RECTS.items():
            x1, y1, x2, y2 = rect
            fill = color if segment in SEGMENTS[digit] else inactive
            draw.rectangle((origin_x + x1, top + y1, origin_x + x2, top + y2), fill=fill)
    draw.ellipse((left + 168, top + 37, left + 174, top + 43), fill=color)


def create_demo_hmi_png(*, low_confidence: bool = True) -> bytes:
    image = Image.new("RGB", IMAGE_SIZE, (18, 31, 37))
    draw = ImageDraw.Draw(image)
    draw.rectangle((25, 25, 775, 475), outline=(83, 121, 128), width=3)
    draw.text((55, 50), "DEMO INJECTION MOLDING HMI", fill=(215, 232, 231))
    for spec in FIELD_SPECS:
        x, y, w, h = spec.region
        box = (int(x * 800), int(y * 500), int((x + w) * 800), int((y + h) * 500))
        draw.text((60, box[1] + 15), f"{spec.label} [{spec.unit}]", fill=(171, 200, 200))
        draw.rectangle(box, fill=(7, 18, 16), outline=(62, 103, 86), width=2)
        _draw_number(
            draw,
            box,
            spec.demo_value,
            dim=low_confidence and spec.code == "holding_pressure_mpa",
        )
    output = io.BytesIO()
    image.save(output, format="PNG", optimize=True)
    return output.getvalue()


def _sample(image: Image.Image, x: int, y: int) -> float:
    pixels = []
    for offset_x in range(-1, 2):
        for offset_y in range(-1, 2):
            pixels.append(image.getpixel((x + offset_x, y + offset_y)))
    return sum(pixels) / len(pixels)


def _recognize_region(region: Image.Image) -> tuple[float | None, str, float]:
    gray = ImageOps.grayscale(region.resize(REGION_SIZE))
    digits: list[str] = []
    confidences: list[float] = []
    reverse = {"".join(sorted(value)): key for key, value in SEGMENTS.items()}
    for origin_x in DIGIT_ORIGINS:
        active: list[str] = []
        samples: list[float] = []
        for segment, point in SEGMENT_POINTS.items():
            intensity = _sample(gray, origin_x + point[0], point[1])
            samples.append(intensity)
            if intensity >= 80:
                active.append(segment)
        digit = reverse.get("".join(sorted(active)))
        if digit is None:
            return None, "", 0.0
        digits.append(digit)
        for value in samples:
            if value >= 80:
                confidences.append(min(1.0, (value - 80) / 80))
            else:
                confidences.append(min(1.0, (80 - value) / 45))
    raw = f"{int(''.join(digits[:3]))}.{digits[3]}"
    return float(raw), raw, round(sum(confidences) / len(confidences), 4)


def _validate_image(data: bytes, filename: str) -> tuple[Image.Image, str]:
    if len(data) > settings.MAX_HMI_UPLOAD_BYTES:
        raise HMIValidationError("VALIDATION_FILE_TOO_LARGE", "HMI image exceeds the 10 MB limit.")
    try:
        source = Image.open(io.BytesIO(data))
        detected = source.format
        source.verify()
        source = Image.open(io.BytesIO(data))
        source = ImageOps.exif_transpose(source).convert("RGB")
    except (UnidentifiedImageError, OSError, Image.DecompressionBombError) as exc:
        raise HMIValidationError(
            "OCR_INVALID_IMAGE", "The file is not a safe readable image."
        ) from exc
    if detected not in {"PNG", "JPEG"}:
        raise HMIValidationError(
            "VALIDATION_UNSUPPORTED_FORMAT", "Only PNG and JPEG are supported."
        )
    if source.width * source.height > MAX_IMAGE_PIXELS:
        raise HMIValidationError("VALIDATION_IMAGE_DIMENSIONS", "Image dimensions are too large.")
    expected_extension = "png" if detected == "PNG" else "jpg"
    if filename.rsplit(".", 1)[-1].lower() not in {"png", "jpg", "jpeg"}:
        raise HMIValidationError(
            "VALIDATION_UNSUPPORTED_FORMAT", "Filename must end in PNG or JPG."
        )
    return source, expected_extension


@transaction.atomic
def create_hmi_extraction(
    upload, *, profile: str = f"{PROFILE_KEY}@{PROFILE_VERSION}"
) -> HMIExtraction:
    try:
        profile_key, profile_version = profile.rsplit("@", 1)
    except ValueError as exc:
        raise HMIValidationError(
            "VALIDATION_HMI_PROFILE", "HMI profile must use the key@version format."
        ) from exc
    expected_checksum = _sha256(json.dumps(_profile_specs_payload(), sort_keys=True).encode())
    profile_definition = HMIProfileVersion.objects.filter(
        profile_key=profile_key,
        version=profile_version,
        status=HMIProfileVersion.Status.PUBLISHED,
    ).first()
    if profile_definition is None or profile_definition.profile_checksum != expected_checksum:
        raise HMIValidationError(
            "VALIDATION_HMI_PROFILE",
            "The selected HMI profile is not published or is incompatible with this extractor.",
        )
    data = upload.read()
    image, extension = _validate_image(data, upload.name)
    digest = _sha256(data)
    artifact = Artifact.objects.create(
        name=upload.name,
        kind=Artifact.Kind.HMI_SOURCE,
        dataset_id="synthetic-hmi-demo-v1",
        classification="public_demo",
    )
    version_id = uuid.uuid4()
    storage_key = f"hmi/source/{digest[:2]}/{version_id}.{extension}"
    default_storage.save(storage_key, ContentFile(data))
    version = ArtifactVersion.objects.create(
        id=version_id,
        artifact=artifact,
        original_filename=upload.name,
        media_type="image/png" if extension == "png" else "image/jpeg",
        format=extension,
        size_bytes=len(data),
        sha256=digest,
        storage_key=storage_key,
        source_system="hmi_upload",
        classification="public_demo",
        malware_status=ArtifactVersion.MalwareStatus.BASIC_SCREENED,
    )
    extraction = HMIExtraction.objects.create(
        image_artifact_version=version,
        profile_definition=profile_definition,
        profile_key=profile_key,
        profile_version=profile_version,
        extractor_version=EXTRACTOR_VERSION,
        status=HMIExtraction.Status.SUCCEEDED,
        image_width=image.width,
        image_height=image.height,
        preprocessing={
            "exif_orientation_applied": True,
            "rotation_correction_degrees": 0,
            "perspective_normalization": "profile_region_normalization",
            "arbitrary_perspective_supported": False,
        },
    )
    needs_review = False
    for spec in FIELD_SPECS:
        x, y, w, h = spec.region
        crop = image.crop(
            (
                int(x * image.width),
                int(y * image.height),
                int((x + w) * image.width),
                int((y + h) * image.height),
            )
        )
        value, raw, confidence = _recognize_region(crop)
        in_range = value is not None and spec.minimum <= value <= spec.maximum
        validation = "valid" if in_range else "unreadable_or_out_of_range"
        review_status = HMIExtractedField.ReviewStatus.NOT_REQUIRED
        if confidence < CONFIDENCE_THRESHOLD or not in_range:
            review_status = HMIExtractedField.ReviewStatus.NEEDS_REVIEW
            needs_review = True
        HMIExtractedField.objects.create(
            extraction=extraction,
            parameter_code=spec.code,
            display_label=spec.label,
            raw_text=f"{raw} {spec.unit}" if raw else "",
            normalized_value=value,
            unit=spec.unit,
            confidence=confidence,
            source_region={"x": x, "y": y, "w": w, "h": h, "coordinate_space": "normalized"},
            validation_status=validation,
            review_status=review_status,
        )
    extraction.review_status = "needs_review" if needs_review else "ready_for_export"
    extraction.save(update_fields=["review_status", "updated_at"])
    return extraction


def field_payload(field: HMIExtractedField) -> dict[str, object]:
    effective_value = (
        field.reviewer_value if field.reviewer_value is not None else field.normalized_value
    )
    effective_unit = field.reviewer_unit or field.unit
    return {
        "field_id": str(field.id),
        "parameter_code": field.parameter_code,
        "display_label": field.display_label,
        "raw_text": field.raw_text,
        "value": field.normalized_value,
        "unit": field.unit,
        "confidence": field.confidence,
        "source_region": field.source_region,
        "validation_status": field.validation_status,
        "review_status": field.review_status,
        "reviewer_correction": (
            {
                "value": field.reviewer_value,
                "unit": field.reviewer_unit,
                "reviewed_by": field.reviewed_by,
            }
            if field.reviewed_at
            else None
        ),
        "effective_value": effective_value,
        "effective_unit": effective_unit,
        "correction_decisions": [
            {
                "decision_id": str(item.id),
                "action": item.action,
                "before_value": item.before_value,
                "after_value": item.after_value,
                "reason": item.reason,
                "decided_by": item.decided_by,
                "created_at": item.created_at.isoformat(),
            }
            for item in field.correction_decisions.all()
        ],
    }


def extraction_payload(extraction: HMIExtraction) -> dict[str, object]:
    fields = [field_payload(field) for field in extraction.fields.all()]
    exports = [
        {
            "export_id": str(item.id),
            "artifact_version_id": str(item.artifact_version_id),
            "template_version": item.template_version,
            "download_url": f"/api/v1/artifact-versions/{item.artifact_version_id}/download",
            "created_at": item.created_at.isoformat(),
        }
        for item in extraction.exports.select_related("artifact_version").all()
    ]
    return {
        "schema_version": "1.0",
        "extraction_id": str(extraction.id),
        "image_artifact_version_id": str(extraction.image_artifact_version_id),
        "image_sha256": extraction.image_artifact_version.sha256,
        "image_download_url": (
            f"/api/v1/artifact-versions/{extraction.image_artifact_version_id}/download"
        ),
        "profile": f"{extraction.profile_key}@{extraction.profile_version}",
        "profile_definition_id": (
            str(extraction.profile_definition_id) if extraction.profile_definition_id else None
        ),
        "extractor_version": extraction.extractor_version,
        "status": extraction.status,
        "image_dimensions": {"width": extraction.image_width, "height": extraction.image_height},
        "preprocessing": extraction.preprocessing,
        "fields": fields,
        "review_status": extraction.review_status,
        "export_status": "ready"
        if extraction.review_status == "ready_for_export"
        else "blocked_pending_review",
        "exports": exports,
        "lineage_ref": f"hmi-extraction:{extraction.id}",
        "created_at": extraction.created_at.isoformat(),
        "limitations": [
            "Stage 9 recognizes only the fixed synthetic seven-segment Demo HMI profile.",
            "Arbitrary perspective, layouts, fonts, languages, and real-machine screens "
            "require approved profiles.",
            "No image content is sent to a cloud vision provider.",
        ],
    }


@transaction.atomic
def review_hmi_fields(
    extraction: HMIExtraction, decisions: object, *, reviewer: str
) -> HMIExtraction:
    if not isinstance(decisions, list) or not decisions:
        raise HMIValidationError("VALIDATION_REVIEW_FIELDS", "fields must be a non-empty array.")
    spec_by_code = {spec.code: spec for spec in FIELD_SPECS}
    for decision in decisions:
        if not isinstance(decision, dict):
            raise HMIValidationError("VALIDATION_REVIEW_FIELDS", "Each decision must be an object.")
        try:
            field = extraction.fields.get(pk=uuid.UUID(str(decision.get("field_id", ""))))
        except (ValueError, HMIExtractedField.DoesNotExist) as exc:
            raise HMIValidationError(
                "VALIDATION_REVIEW_FIELD_ID", "A field ID is invalid."
            ) from exc
        action = str(decision.get("action", ""))
        before_value = {
            "review_status": field.review_status,
            "reviewer_value": field.reviewer_value,
            "reviewer_unit": field.reviewer_unit,
        }
        if action == "confirm":
            if field.normalized_value is None:
                raise HMIValidationError(
                    "VALIDATION_REVIEW_VALUE", "Unreadable fields require correction or rejection."
                )
            field.review_status = HMIExtractedField.ReviewStatus.CONFIRMED
        elif action == "correct":
            spec = spec_by_code[field.parameter_code]
            try:
                value = float(decision["value"])
            except (KeyError, TypeError, ValueError) as exc:
                raise HMIValidationError(
                    "VALIDATION_REVIEW_VALUE", "Correction value must be numeric."
                ) from exc
            if not spec.minimum <= value <= spec.maximum or decision.get("unit") != spec.unit:
                raise HMIValidationError(
                    "VALIDATION_REVIEW_RANGE",
                    "Correction value or unit is outside the approved profile.",
                )
            field.reviewer_value = value
            field.reviewer_unit = spec.unit
            field.review_status = HMIExtractedField.ReviewStatus.CORRECTED
        elif action == "reject":
            field.review_status = HMIExtractedField.ReviewStatus.REJECTED
        else:
            raise HMIValidationError(
                "VALIDATION_REVIEW_ACTION", "action must be confirm, correct, or reject."
            )
        field.reviewed_by = reviewer[:128]
        field.reviewed_at = timezone.now()
        field.save()
        HMICorrectionDecision.objects.create(
            field=field,
            action=action,
            before_value=before_value,
            after_value={
                "review_status": field.review_status,
                "reviewer_value": field.reviewer_value,
                "reviewer_unit": field.reviewer_unit,
            },
            reason=str(decision.get("reason", "Reviewed extracted value"))[:512],
            decided_by=reviewer[:128],
        )
    unresolved = extraction.fields.filter(
        review_status=HMIExtractedField.ReviewStatus.NEEDS_REVIEW
    ).exists()
    rejected = extraction.fields.filter(
        review_status=HMIExtractedField.ReviewStatus.REJECTED
    ).exists()
    extraction.review_status = (
        "rejected" if rejected else ("needs_review" if unresolved else "ready_for_export")
    )
    extraction.save(update_fields=["review_status", "updated_at"])
    AuditEvent.objects.create(
        event_type="hmi.fields_reviewed.v1",
        actor_id=reviewer[:128],
        target_refs=[f"hmi-extraction:{extraction.id}"],
        detail={"review_status": extraction.review_status, "decision_count": len(decisions)},
        payload_hash=_sha256(json.dumps(decisions, sort_keys=True).encode()),
    )
    return extraction


def _build_workbook(extraction: HMIExtraction) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Parameters"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A9"
    navy = "173B3F"
    teal = "2F766F"
    light = "EAF4F2"
    warning = "FFF2CC"
    thin = Side(style="thin", color="CBD8D7")
    sheet.merge_cells("A1:I1")
    sheet["A1"] = "Mold AI Platform - Reviewed HMI Parameters"
    sheet["A1"].font = Font(bold=True, color="FFFFFF", size=16)
    sheet["A1"].fill = PatternFill("solid", fgColor=navy)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28
    metadata = [
        ("Extraction ID", str(extraction.id)),
        ("Source image SHA-256", extraction.image_artifact_version.sha256),
        ("Profile", f"{extraction.profile_key}@{extraction.profile_version}"),
        ("Extracted at", extraction.created_at.replace(tzinfo=None)),
    ]
    for row, (label, value) in enumerate(metadata, start=2):
        sheet.cell(row, 1, label).font = Font(bold=True, color=navy)
        sheet.cell(row, 2, value)
    sheet["D2"] = "Field count"
    sheet["E2"] = "=COUNTA(A9:A100)"
    sheet["D3"] = "Human-reviewed count"
    sheet["E3"] = '=COUNTIF(H9:H100,"confirmed")+COUNTIF(H9:H100,"corrected")'
    for cell in (sheet["D2"], sheet["D3"]):
        cell.font = Font(bold=True, color=navy)
    headers = [
        "Parameter Code",
        "Display Label",
        "Raw OCR",
        "Value",
        "Unit",
        "Confidence",
        "Source Region",
        "Review Status",
        "Reviewed By",
    ]
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(8, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=teal)
        cell.alignment = Alignment(horizontal="center")
    fields = list(extraction.fields.all())
    for row, field in enumerate(fields, start=9):
        payload = field_payload(field)
        values = [
            field.parameter_code,
            field.display_label,
            field.raw_text,
            payload["effective_value"],
            payload["effective_unit"],
            field.confidence,
            json.dumps(field.source_region, sort_keys=True),
            field.review_status,
            field.reviewed_by,
        ]
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row, column, value)
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=column in {3, 7})
        sheet.cell(row, 4).number_format = "0.000"
        sheet.cell(row, 6).number_format = "0.0%"
        if field.review_status in {"confirmed", "corrected"}:
            sheet.cell(row, 8).fill = PatternFill("solid", fgColor=light)
        elif field.review_status == "needs_review":
            sheet.cell(row, 8).fill = PatternFill("solid", fgColor=warning)
    end_row = 8 + len(fields)
    table = Table(displayName="ReviewedHMIParameters", ref=f"A8:I{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False
    )
    sheet.add_table(table)
    widths = {"A": 28, "B": 24, "C": 18, "D": 12, "E": 12, "F": 14, "G": 48, "H": 18, "I": 20}
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    audit = workbook.create_sheet("Audit")
    audit.sheet_view.showGridLines = False
    audit["A1"] = "Export lineage and limitations"
    audit["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    audit["A1"].fill = PatternFill("solid", fgColor=navy)
    audit.merge_cells("A1:B1")
    audit_rows = [
        ("Template version", TEMPLATE_VERSION),
        ("Image artifact version", str(extraction.image_artifact_version_id)),
        ("Image SHA-256", extraction.image_artifact_version.sha256),
        ("Extractor version", extraction.extractor_version),
        ("Preprocessing", json.dumps(extraction.preprocessing, sort_keys=True)),
        ("Classification", extraction.image_artifact_version.classification),
        ("Notice", "Reviewed extraction only; this workbook does not reproduce the HMI layout."),
    ]
    for row, values in enumerate(audit_rows, start=3):
        audit.cell(row, 1, values[0]).font = Font(bold=True, color=navy)
        audit.cell(row, 2, values[1]).alignment = Alignment(wrap_text=True, vertical="top")
    audit.column_dimensions["A"].width = 28
    audit.column_dimensions["B"].width = 90
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    verified = load_workbook(output, data_only=False, read_only=True)
    if verified.sheetnames != ["Parameters", "Audit"] or verified["Parameters"]["A1"].value is None:
        raise HMIValidationError(
            "XLSX_VERIFICATION_FAILED", "Generated workbook verification failed."
        )
    output.seek(0)
    return output.getvalue()


@transaction.atomic
def export_hmi_workbook(extraction: HMIExtraction, *, created_by: str) -> HMIExport:
    if extraction.review_status != "ready_for_export":
        raise HMIValidationError(
            "CONFLICT_REVIEW_REQUIRED", "Resolve every low-confidence field before export."
        )
    snapshot = [field_payload(field) for field in extraction.fields.all()]
    snapshot_hash = _sha256(json.dumps(snapshot, sort_keys=True).encode())
    data = _build_workbook(extraction)
    artifact = Artifact.objects.create(
        name=f"hmi-parameters-{extraction.id}.xlsx",
        kind=Artifact.Kind.HMI_EXPORT,
        dataset_id=extraction.image_artifact_version.artifact.dataset_id,
        classification=extraction.image_artifact_version.classification,
        created_by=created_by[:128],
    )
    version_id = uuid.uuid4()
    digest = _sha256(data)
    storage_key = f"hmi/export/{digest[:2]}/{version_id}.xlsx"
    default_storage.save(storage_key, ContentFile(data))
    version = ArtifactVersion.objects.create(
        id=version_id,
        artifact=artifact,
        original_filename=artifact.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        format="xlsx",
        size_bytes=len(data),
        sha256=digest,
        storage_key=storage_key,
        source_system="hmi_xlsx_export",
        classification=artifact.classification,
        malware_status=ArtifactVersion.MalwareStatus.BASIC_SCREENED,
    )
    export = HMIExport.objects.create(
        extraction=extraction,
        artifact_version=version,
        template_version=TEMPLATE_VERSION,
        reviewed_snapshot_hash=snapshot_hash,
        created_by=created_by[:128],
    )
    AuditEvent.objects.create(
        event_type="hmi.xlsx_exported.v1",
        actor_id=created_by[:128],
        target_refs=[f"hmi-extraction:{extraction.id}", f"artifact-version:{version.id}"],
        detail={"template_version": TEMPLATE_VERSION, "reviewed_snapshot_hash": snapshot_hash},
        payload_hash=digest,
    )
    return export
