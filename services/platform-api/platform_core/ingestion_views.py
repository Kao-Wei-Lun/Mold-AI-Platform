from __future__ import annotations

import csv
import io
import json

from django.db import IntegrityError, transaction
from django.http import HttpResponse
from openpyxl import load_workbook
from rest_framework.authentication import SessionAuthentication
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.request import Request
from rest_framework.response import Response
from rest_framework.views import APIView

from .identity import audit_identity_event
from .ingestion_adapters import MAX_BATCH_RECORDS, SUPPORTED_INGESTION_DOMAINS
from .ingestion_center import (
    IngestionError,
    attach_source_bytes,
    ensure_inline_source,
    persist_validation,
)
from .models import BulkImportBatch, DataScope, Job, JobEvent
from .pagination import PaginationValueError, paginate
from .tasks import commit_ingestion_job

TEMPLATE_HEADERS = {
    "master_data": ["kind", "code", "name_en", "name_zh_tw"],
    "projects": ["code", "name", "description"],
    "registry": [
        "project_code",
        "project_name",
        "part_number",
        "part_name",
        "product_type",
        "material_code",
        "mold_code",
        "mold_name",
        "mold_type",
        "cavity_count",
        "revision_code",
        "change_summary",
    ],
    "rule_profiles": [
        "profile_key",
        "version",
        "rule_id",
        "title",
        "description",
        "evaluator",
        "operator",
        "limit_value",
        "unit",
        "tolerance",
        "severity",
        "risk_type",
        "recommendation",
        "reference_document",
        "reference_revision",
        "mold_type",
        "product_type",
        "material",
        "molding_process",
    ],
    "trials": [
        "case_code",
        "mold_revision_ref",
        "part_revision_ref",
        "machine_code",
        "material_code",
        "material_lot",
        "product_type",
        "purpose",
        "outcome",
        "started_at",
        "run_number",
        "result",
        "parameter_code",
        "parameter_name",
        "parameter_value",
        "parameter_unit",
        "value_kind",
        "sampling_method",
    ],
}
SUPPORTED_SOURCE_SUFFIXES = {"json", "csv", "xlsx"}


def _error(request: Request, code: str, message: str, status: int) -> Response:
    return Response(
        {
            "error": {
                "code": code,
                "message": message,
                "retryable": status >= 500,
                "request_id": getattr(request._request, "mold_ai_request_id", ""),
            }
        },
        status=status,
    )


def _require_any(request: Request, *permissions: str) -> Response | None:
    granted = getattr(request._request, "mold_ai_permissions", set())
    if any(permission in granted for permission in permissions):
        return None
    return _error(request, "ACCESS_DENIED", f"The account does not grant {permissions[0]}.", 403)


def _actor(request: Request) -> str:
    return str(getattr(request._request, "mold_ai_actor_id", "anonymous"))


def _scope(
    request: Request, *, source: dict | None = None
) -> tuple[DataScope | None, Response | None]:
    code = str((source or {}).get("scope") or request.query_params.get("scope", "public-demo"))
    allowed = set(getattr(request._request, "mold_ai_data_scopes", set()))
    if code not in allowed:
        return None, _error(request, "SCOPE_ACCESS_DENIED", "The data scope is not assigned.", 403)
    scope = DataScope.objects.filter(code=code, is_active=True).first()
    if scope is None:
        return None, _error(request, "SCOPE_NOT_FOUND", "The data scope does not exist.", 404)
    return scope, None


def _batch(request: Request, batch_id: str) -> tuple[BulkImportBatch | None, Response | None]:
    batch = (
        BulkImportBatch.objects.select_related("scope", "job")
        .prefetch_related("source_files", "issues", "record_results")
        .filter(id=batch_id)
        .first()
    )
    allowed = set(getattr(request._request, "mold_ai_data_scopes", set()))
    if batch is None or batch.scope.code not in allowed:
        return None, _error(request, "NOT_FOUND", "Ingestion batch not found.", 404)
    return batch, None


def _batch_payload(
    request: Request, batch: BulkImportBatch, *, detail: bool = True
) -> dict[str, object]:
    payload: dict[str, object] = {
        "schema_version": "1.0",
        "batch_id": str(batch.id),
        "canonical_id": f"ingestion:{batch.id}",
        "scope": batch.scope.code,
        "classification": batch.classification,
        "domain": batch.domain,
        "source_name": batch.source_name,
        "status": batch.status,
        "mapping_version": batch.mapping_version,
        "validation": batch.validation_result,
        "reconciliation": batch.reconciliation,
        "job_id": str(batch.job_id) if batch.job_id else None,
        "created_by": batch.created_by,
        "created_at": batch.created_at.isoformat(),
        "updated_at": batch.updated_at.isoformat(),
        "deep_link": f"/data/imports/{batch.id}",
        "request_id": getattr(request._request, "mold_ai_request_id", ""),
        "correlation_id": str(batch.job.correlation_id) if batch.job_id else None,
    }
    if detail:
        payload.update(
            {
                "field_mapping": batch.field_mapping,
                "records": batch.records,
                "source_files": [
                    {
                        "source_file_id": str(item.id),
                        "artifact_version_id": str(item.artifact_version_id),
                        "file_name": item.file_name,
                        "sha256": item.sha256,
                        "mime_type": item.mime_type,
                        "size_bytes": item.size_bytes,
                        "screening": item.screening,
                    }
                    for item in batch.source_files.all()
                ],
                "issues": [
                    {
                        "issue_id": str(item.id),
                        "row_number": item.row_number,
                        "field_name": item.field_name,
                        "code": item.code,
                        "message": item.message,
                        "raw_value": item.raw_value,
                        "suggestion": item.suggestion,
                        "severity": item.severity,
                    }
                    for item in batch.issues.all()
                ],
            }
        )
    return payload


def _parse_source(file_name: str, content: bytes) -> list[dict[str, object]]:
    suffix = file_name.rsplit(".", maxsplit=1)[-1].lower() if "." in file_name else ""
    if suffix not in SUPPORTED_SOURCE_SUFFIXES:
        raise IngestionError(
            "IMPORT_SOURCE_FORMAT_UNSUPPORTED",
            "The ingestion foundation accepts JSON, CSV, or XLSX.",
        )
    try:
        if suffix == "json":
            parsed = json.loads(content.decode("utf-8-sig"))
            if not isinstance(parsed, list):
                raise ValueError("JSON root must be an array.")
            records = parsed
        elif suffix == "csv":
            records = list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))
        else:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            headers = [str(value or "").strip() for value in (rows[0] if rows else [])]
            records = [
                {headers[index]: value for index, value in enumerate(row) if headers[index]}
                for row in rows[1:]
                if any(value is not None for value in row)
            ]
    except (UnicodeDecodeError, json.JSONDecodeError, ValueError) as exc:
        raise IngestionError("IMPORT_SOURCE_PARSE_FAILED", str(exc)) from exc
    if not all(isinstance(item, dict) for item in records):
        raise IngestionError("IMPORT_SOURCE_SCHEMA_INVALID", "Every source row must be an object.")
    if not 1 <= len(records) <= MAX_BATCH_RECORDS:
        raise IngestionError(
            "IMPORT_BATCH_SIZE", f"A source must contain 1–{MAX_BATCH_RECORDS} records."
        )
    return records


class IngestionListCreateView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes: list = []

    def get(self, request: Request) -> Response:
        if denied := _require_any(request, "ingestion:read", "enterprise:read"):
            return denied
        scope, invalid = _scope(request)
        if invalid:
            return invalid
        batches = BulkImportBatch.objects.select_related("scope", "job").filter(scope=scope)
        if status := request.query_params.get("status"):
            batches = batches.filter(status=status)
        try:
            records, page = paginate(
                request,
                batches,
                allowed_sort={"created_at": "created_at", "status": "status", "domain": "domain"},
                default_sort="-created_at",
            )
        except PaginationValueError as exc:
            return _error(request, "VALIDATION_PAGINATION", str(exc), 400)
        return Response(
            {
                "schema_version": "1.0",
                "items": [_batch_payload(request, item, detail=False) for item in records],
                "page": page,
            }
        )

    def post(self, request: Request) -> Response:
        if denied := _require_any(request, "ingestion:create", "bulk:manage"):
            return denied
        scope, invalid = _scope(request, source=request.data)
        if invalid:
            return invalid
        domain = str(request.data.get("domain", ""))
        if domain not in SUPPORTED_INGESTION_DOMAINS:
            return _error(request, "IMPORT_DOMAIN_UNSUPPORTED", "Unsupported import domain.", 400)
        key = str(request.data.get("idempotency_key", "")).strip()
        if not key:
            return _error(
                request, "VALIDATION_IDEMPOTENCY_KEY", "idempotency_key is required.", 400
            )
        existing = (
            BulkImportBatch.objects.select_related("scope", "job")
            .filter(idempotency_key=key[:255])
            .first()
        )
        if existing:
            if existing.scope_id != scope.id or existing.domain != domain:
                return _error(
                    request,
                    "IDEMPOTENCY_SCOPE_CONFLICT",
                    "The key belongs to another scope or domain.",
                    409,
                )
            return Response(_batch_payload(request, existing), status=200)
        records = request.data.get("records", [])
        mapping = request.data.get("field_mapping", {})
        if not isinstance(records, list) or len(records) > MAX_BATCH_RECORDS:
            return _error(
                request,
                "IMPORT_BATCH_SIZE",
                f"A batch supports at most {MAX_BATCH_RECORDS} records.",
                400,
            )
        if not isinstance(mapping, dict):
            return _error(
                request, "IMPORT_MAPPING_INVALID", "field_mapping must be an object.", 400
            )
        try:
            batch = BulkImportBatch.objects.create(
                scope=scope,
                domain=domain,
                source_name=str(request.data.get("source_name", "manual"))[:255],
                idempotency_key=key[:255],
                schema_version=str(request.data.get("schema_version", "1.0"))[:32],
                classification=scope.classification,
                records=records,
                field_mapping=mapping,
                status=BulkImportBatch.Status.DRAFT,
                created_by=_actor(request),
            )
        except IntegrityError:
            return _error(request, "IDEMPOTENCY_CONFLICT", "The ingestion key already exists.", 409)
        audit_identity_event(
            "ingestion.created.v1",
            actor_id=_actor(request),
            target_refs=[f"ingestion:{batch.id}", f"scope:{scope.code}"],
            detail={"domain": domain, "record_count": len(records)},
        )
        return Response(_batch_payload(request, batch), status=201)


class IngestionDetailView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes: list = []

    def get(self, request: Request, batch_id: str) -> Response:
        if denied := _require_any(request, "ingestion:read", "enterprise:read"):
            return denied
        batch, invalid = _batch(request, batch_id)
        return invalid or Response(_batch_payload(request, batch))


class IngestionFileView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes: list = []
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request: Request, batch_id: str) -> Response:
        if denied := _require_any(request, "ingestion:create", "bulk:manage"):
            return denied
        batch, invalid = _batch(request, batch_id)
        if invalid:
            return invalid
        if batch.status not in {BulkImportBatch.Status.DRAFT, BulkImportBatch.Status.UPLOADED}:
            return _error(
                request, "INGESTION_STATE_CONFLICT", "Files cannot be added in this state.", 409
            )
        upload = request.FILES.get("file")
        if upload is None:
            return _error(request, "VALIDATION_FILE_REQUIRED", "A source file is required.", 400)
        content = upload.read()
        try:
            records = _parse_source(upload.name, content)
            attach_source_bytes(
                batch,
                file_name=upload.name,
                content=content,
                mime_type=upload.content_type or "application/octet-stream",
            )
        except IngestionError as exc:
            return _error(request, exc.code, exc.user_message, 400)
        batch.records = records
        batch._prefetched_objects_cache.pop("source_files", None)
        batch.source_name = upload.name[:255]
        batch.status = (
            BulkImportBatch.Status.UPLOADED
            if batch.field_mapping
            else BulkImportBatch.Status.MAPPING_REQUIRED
        )
        batch.save(update_fields=["records", "source_name", "status", "updated_at"])
        return Response(_batch_payload(request, batch), status=201)


class IngestionMappingView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes: list = []
    parser_classes = [JSONParser]

    def put(self, request: Request, batch_id: str) -> Response:
        if denied := _require_any(request, "ingestion:create", "bulk:manage"):
            return denied
        batch, invalid = _batch(request, batch_id)
        if invalid:
            return invalid
        mapping = request.data.get("field_mapping", {})
        if not isinstance(mapping, dict):
            return _error(
                request, "IMPORT_MAPPING_INVALID", "field_mapping must be an object.", 400
            )
        batch.field_mapping = mapping
        batch.mapping_version = str(request.data.get("mapping_version", "1.0"))[:32]
        batch.status = BulkImportBatch.Status.UPLOADED
        batch.save(update_fields=["field_mapping", "mapping_version", "status", "updated_at"])
        return Response(_batch_payload(request, batch))


class IngestionValidateView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes: list = []

    def post(self, request: Request, batch_id: str) -> Response:
        if denied := _require_any(request, "ingestion:validate", "bulk:manage"):
            return denied
        batch, invalid = _batch(request, batch_id)
        if invalid:
            return invalid
        if batch.status in {BulkImportBatch.Status.COMMITTED, BulkImportBatch.Status.COMMITTING}:
            return _error(
                request, "INGESTION_STATE_CONFLICT", "Committed data cannot be revalidated.", 409
            )
        if not batch.records:
            return _error(
                request, "IMPORT_SOURCE_REQUIRED", "Upload a source or provide records first.", 409
            )
        persist_validation(batch)
        batch.refresh_from_db()
        return Response(_batch_payload(request, batch))


class IngestionIssuesView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes: list = []

    def get(self, request: Request, batch_id: str) -> Response:
        if denied := _require_any(request, "ingestion:read", "enterprise:read"):
            return denied
        batch, invalid = _batch(request, batch_id)
        if invalid:
            return invalid
        payload = _batch_payload(request, batch)
        return Response(
            {"schema_version": "1.0", "batch_id": str(batch.id), "items": payload["issues"]}
        )


class IngestionCommitView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes: list = []

    def post(self, request: Request, batch_id: str) -> Response:
        if denied := _require_any(request, "ingestion:commit", "bulk:manage"):
            return denied
        batch, invalid = _batch(request, batch_id)
        if invalid:
            return invalid
        if batch.status == BulkImportBatch.Status.COMMITTED:
            return Response(_batch_payload(request, batch), status=200)
        if batch.status in {BulkImportBatch.Status.QUEUED, BulkImportBatch.Status.COMMITTING}:
            return Response(_batch_payload(request, batch), status=202)
        if batch.status != BulkImportBatch.Status.VALIDATED:
            return _error(request, "IMPORT_NOT_COMMITTABLE", "A passing dry run is required.", 409)
        reason = str(request.data.get("reason", "")).strip()
        if not reason:
            return _error(request, "VALIDATION_REASON_REQUIRED", "A reason is required.", 400)
        version = ensure_inline_source(batch)
        with transaction.atomic():
            job = Job.objects.create(
                capability_id="data.ingestion.commit",
                capability_version="1.0.0",
                state=Job.State.QUEUED,
                queue="general",
                resource_class="cpu",
                input_artifact_version=version,
                input_snapshot={
                    "schema_version": batch.schema_version,
                    "batch_id": str(batch.id),
                    "source_sha256": version.sha256,
                    "mapping_version": batch.mapping_version,
                    "reason": reason[:512],
                },
                idempotency_key=f"ingestion-commit:{batch.id}",
            )
            JobEvent.objects.create(
                job=job,
                from_state="",
                to_state=Job.State.QUEUED,
                stage="queued",
                progress=0,
            )
            batch.job = job
            batch.status = BulkImportBatch.Status.QUEUED
            batch.save(update_fields=["job", "status", "updated_at"])
            transaction.on_commit(
                lambda: commit_ingestion_job.apply_async(
                    args=[str(job.id), str(batch.id), _actor(request)], queue="general"
                )
            )
        return Response(_batch_payload(request, batch), status=202)


class IngestionCancelView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes: list = []

    def post(self, request: Request, batch_id: str) -> Response:
        if denied := _require_any(request, "ingestion:cancel", "bulk:manage"):
            return denied
        batch, invalid = _batch(request, batch_id)
        if invalid:
            return invalid
        if batch.status in {BulkImportBatch.Status.COMMITTED, BulkImportBatch.Status.COMMITTING}:
            return _error(
                request,
                "INGESTION_STATE_CONFLICT",
                "This ingestion can no longer be cancelled.",
                409,
            )
        batch.status = BulkImportBatch.Status.CANCELLED
        batch.save(update_fields=["status", "updated_at"])
        if batch.job_id:
            Job.objects.filter(id=batch.job_id, state=Job.State.QUEUED).update(
                state=Job.State.CANCELLED, stage="cancelled", progress=100
            )
        return Response(_batch_payload(request, batch))


class IngestionReconciliationView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes: list = []

    def get(self, request: Request, batch_id: str) -> Response:
        if denied := _require_any(request, "ingestion:read", "enterprise:read"):
            return denied
        batch, invalid = _batch(request, batch_id)
        if invalid:
            return invalid
        return Response(
            {
                "schema_version": "1.0",
                "batch_id": str(batch.id),
                "status": batch.status,
                "reconciliation": batch.reconciliation,
            }
        )


class ImportTemplateView(APIView):
    authentication_classes = [SessionAuthentication]
    permission_classes: list = []

    def get(self, request: Request, domain: str) -> HttpResponse | Response:
        if denied := _require_any(request, "ingestion:read", "enterprise:read"):
            return denied
        headers = TEMPLATE_HEADERS.get(domain)
        if headers is None:
            return _error(request, "IMPORT_DOMAIN_UNSUPPORTED", "No template is available.", 404)
        stream = io.StringIO()
        writer = csv.writer(stream, lineterminator="\n")
        writer.writerow(headers)
        response = HttpResponse(stream.getvalue(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{domain}-template-v1.csv"'
        response["X-Schema-Version"] = "1.0"
        response["X-Request-ID"] = getattr(request._request, "mold_ai_request_id", "")
        return response
