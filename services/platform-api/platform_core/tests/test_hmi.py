import io
from tempfile import TemporaryDirectory

from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from openpyxl import load_workbook
from PIL import Image

from platform_core.hmi import FIELD_SPECS, create_demo_hmi_png
from platform_core.models import Artifact, AuditEvent, HMIExport, HMIExtraction


class HMIExtractionEndpointTests(TestCase):
    def setUp(self) -> None:
        self.media_directory = TemporaryDirectory()
        self.settings_override = override_settings(MEDIA_ROOT=self.media_directory.name)
        self.settings_override.enable()

    def tearDown(self) -> None:
        self.settings_override.disable()
        self.media_directory.cleanup()

    @staticmethod
    def upload(low_confidence: bool = True, name: str = "demo-hmi.png"):
        return SimpleUploadedFile(
            name,
            create_demo_hmi_png(low_confidence=low_confidence),
            content_type="image/png",
        )

    def create_extraction(self, *, low_confidence: bool = True):
        return self.client.post(
            "/api/v1/hmi-extractions",
            {"file": self.upload(low_confidence), "profile": "demo-generic-injection@1.0"},
        )

    def test_clean_profile_recognizes_all_critical_numeric_fields_exactly(self) -> None:
        response = self.create_extraction(low_confidence=False)

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        self.assertEqual(payload["review_status"], "ready_for_export")
        self.assertEqual(payload["export_status"], "ready")
        self.assertEqual(payload["profile"], "demo-generic-injection@1.0")
        expected = {spec.code: spec.demo_value for spec in FIELD_SPECS}
        actual = {field["parameter_code"]: field["value"] for field in payload["fields"]}
        self.assertEqual(actual, expected)
        self.assertTrue(all(field["confidence"] >= 0.95 for field in payload["fields"]))
        self.assertTrue(
            all(field["review_status"] == "not_required" for field in payload["fields"])
        )
        self.assertEqual(HMIExtraction.objects.count(), 1)
        self.assertEqual(Artifact.objects.filter(kind=Artifact.Kind.HMI_SOURCE).count(), 1)

        detail = self.client.get(f"/api/v1/hmi-extractions/{payload['extraction_id']}")
        listing = self.client.get("/api/v1/hmi-extractions")
        self.assertEqual(detail.status_code, 200)
        self.assertEqual(detail.json()["image_sha256"], payload["image_sha256"])
        self.assertEqual(listing.status_code, 200)
        self.assertEqual(len(listing.json()["items"]), 1)

    def test_low_confidence_field_is_exact_but_blocks_export_until_reviewed(self) -> None:
        response = self.create_extraction()
        self.assertEqual(response.status_code, 201)
        payload = response.json()
        holding = next(
            field
            for field in payload["fields"]
            if field["parameter_code"] == "holding_pressure_mpa"
        )
        self.assertEqual(holding["value"], 55.0)
        self.assertLess(holding["confidence"], 0.90)
        self.assertEqual(holding["review_status"], "needs_review")
        self.assertEqual(payload["review_status"], "needs_review")

        blocked = self.client.post(
            f"/api/v1/hmi-extractions/{payload['extraction_id']}/exports",
            {"created_by": "qa-user"},
            content_type="application/json",
        )
        self.assertEqual(blocked.status_code, 409)
        self.assertEqual(blocked.json()["error"]["code"], "CONFLICT_REVIEW_REQUIRED")
        self.assertEqual(HMIExport.objects.count(), 0)

        reviewed = self.client.post(
            f"/api/v1/hmi-extractions/{payload['extraction_id']}/review",
            {
                "reviewed_by": "qa-user",
                "fields": [{"field_id": holding["field_id"], "action": "confirm"}],
            },
            content_type="application/json",
        )
        self.assertEqual(reviewed.status_code, 200)
        self.assertEqual(reviewed.json()["review_status"], "ready_for_export")
        reviewed_holding = next(
            field
            for field in reviewed.json()["fields"]
            if field["parameter_code"] == "holding_pressure_mpa"
        )
        self.assertEqual(reviewed_holding["review_status"], "confirmed")
        self.assertEqual(reviewed_holding["reviewer_correction"]["reviewed_by"], "qa-user")
        self.assertTrue(AuditEvent.objects.filter(event_type="hmi.fields_reviewed.v1").exists())

    def test_reviewed_export_is_versioned_and_has_valid_typed_workbook(self) -> None:
        payload = self.create_extraction().json()
        holding = next(
            field
            for field in payload["fields"]
            if field["parameter_code"] == "holding_pressure_mpa"
        )
        self.client.post(
            f"/api/v1/hmi-extractions/{payload['extraction_id']}/review",
            {
                "reviewed_by": "engineer-1",
                "fields": [
                    {
                        "field_id": holding["field_id"],
                        "action": "correct",
                        "value": 56,
                        "unit": "MPa",
                    }
                ],
            },
            content_type="application/json",
        )
        exported = self.client.post(
            f"/api/v1/hmi-extractions/{payload['extraction_id']}/exports",
            {"created_by": "engineer-1"},
            content_type="application/json",
        )
        self.assertEqual(exported.status_code, 201)
        export_payload = exported.json()
        self.assertEqual(export_payload["template_version"], "reviewed-parameters@1.0.0")
        self.assertEqual(HMIExport.objects.count(), 1)
        self.assertEqual(Artifact.objects.filter(kind=Artifact.Kind.HMI_EXPORT).count(), 1)

        download = self.client.get(export_payload["download_url"])
        self.assertEqual(download.status_code, 200)
        self.assertIn("attachment", download["Content-Disposition"])
        workbook_bytes = b"".join(download.streaming_content)
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=False)
        self.assertEqual(workbook.sheetnames, ["Parameters", "Audit"])
        parameters = workbook["Parameters"]
        audit = workbook["Audit"]
        self.assertEqual(parameters["A1"].value, "Mold AI Platform - Reviewed HMI Parameters")
        self.assertEqual(parameters["B3"].value, payload["image_sha256"])
        self.assertEqual(parameters["E2"].value, "=COUNTA(A9:A100)")
        values_by_code = {
            parameters.cell(row, 1).value: parameters.cell(row, 4).value for row in range(9, 13)
        }
        self.assertEqual(values_by_code["holding_pressure_mpa"], 56)
        self.assertIsInstance(values_by_code["injection_pressure_mpa"], (int, float))
        self.assertEqual(audit["B5"].value, payload["image_sha256"])
        for row in workbook.worksheets:
            for cells in row.iter_rows():
                for cell in cells:
                    if isinstance(cell.value, str):
                        self.assertNotIn(cell.value, {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?"})
        self.assertTrue(AuditEvent.objects.filter(event_type="hmi.xlsx_exported.v1").exists())

    def test_reject_and_invalid_correction_keep_export_blocked(self) -> None:
        payload = self.create_extraction().json()
        holding = next(
            field for field in payload["fields"] if field["review_status"] == "needs_review"
        )
        endpoint = f"/api/v1/hmi-extractions/{payload['extraction_id']}/review"
        invalid = self.client.post(
            endpoint,
            {
                "fields": [
                    {
                        "field_id": holding["field_id"],
                        "action": "correct",
                        "value": 999,
                        "unit": "bar",
                    }
                ]
            },
            content_type="application/json",
        )
        self.assertEqual(invalid.status_code, 400)
        self.assertEqual(invalid.json()["error"]["code"], "VALIDATION_REVIEW_RANGE")

        rejected = self.client.post(
            endpoint,
            {"fields": [{"field_id": holding["field_id"], "action": "reject"}]},
            content_type="application/json",
        )
        self.assertEqual(rejected.status_code, 200)
        self.assertEqual(rejected.json()["review_status"], "rejected")
        blocked = self.client.post(
            f"/api/v1/hmi-extractions/{payload['extraction_id']}/exports",
            {},
            content_type="application/json",
        )
        self.assertEqual(blocked.status_code, 409)

    def test_fixture_and_upload_validation(self) -> None:
        fixture = self.client.get("/api/v1/hmi/demo-fixture?variant=low-confidence")
        self.assertEqual(fixture.status_code, 200)
        self.assertEqual(fixture["Content-Type"], "image/png")
        image = Image.open(io.BytesIO(fixture.content))
        self.assertEqual(image.size, (800, 500))

        invalid_variant = self.client.get("/api/v1/hmi/demo-fixture?variant=unknown")
        self.assertEqual(invalid_variant.status_code, 400)
        missing = self.client.post("/api/v1/hmi-extractions", {})
        self.assertEqual(missing.status_code, 400)
        bogus = self.client.post(
            "/api/v1/hmi-extractions",
            {"file": SimpleUploadedFile("screen.png", b"not-image", "image/png")},
        )
        self.assertEqual(bogus.status_code, 400)
        self.assertEqual(bogus.json()["error"]["code"], "OCR_INVALID_IMAGE")

    @override_settings(MAX_HMI_UPLOAD_BYTES=8)
    def test_upload_size_limit_is_enforced_before_image_processing(self) -> None:
        response = self.client.post(
            "/api/v1/hmi-extractions",
            {"file": SimpleUploadedFile("screen.png", b"0123456789", "image/png")},
        )
        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["error"]["code"], "VALIDATION_FILE_TOO_LARGE")
